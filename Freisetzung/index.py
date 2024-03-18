import pandas as pd
import numpy as np
import math
from scipy.optimize import curve_fit
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

file_path_blinds = r'e:\Dokumente\Ausbildung\03-Promotion\Messdaten-automatisieren\Freisetzung-06.08.19_CPPA-APPA\saz_Abs_MB_0807_095415_MOPS.xlsx'
file_path_samples = r'e:\Dokumente\Ausbildung\03-Promotion\Messdaten-automatisieren\Freisetzung-06.08.19_CPPA-APPA\saz_Abs_MB_0812_095833_1MOPS-orig.xlsx'
file_path_concentrations = r'e:\Dokumente\Ausbildung\03-Promotion\Messdaten-automatisieren\Freisetzung-06.08.19_CPPA-APPA\konzentrationen.xlsx'
file_path_output = r'e:\Dokumente\Ausbildung\03-Promotion\Messdaten-automatisieren\Freisetzung-06.08.19_CPPA-APPA\result-sheet.xlsx'

# opens two excel sheets (values and blinds) and returns the relevant data after substracting blinds as pandas.DataFrame object
def get_data(file_path_blinds, file_path_samples):
    # Read a specific range of rows (only relevant data from file format)
    values_blinds = pd.read_excel(file_path_blinds, sheet_name='Result sheet', skiprows=55, nrows=98)
    values_samples = pd.read_excel(file_path_samples, sheet_name='Result sheet', skiprows=55, nrows=98)

    # Copy the first two lines that are additional info
    values = values_samples.iloc[:2].copy()

    # Perform substraction of blinds (without first line / A1, A2 ... H11, H12)
    substracted_values = values_samples.iloc[2:, 1:] - values_blinds.iloc[2:, 1:]

    # Set negative values to zero
    substracted_values[substracted_values < 0] = 0

    # Merge the results
    substracted_values = pd.concat([values_samples.iloc[:, :1], substracted_values], axis=1)
    values = pd.concat([values, substracted_values[2:]])

    return values

# opens excel sheet (with previously calculated concentrations) and returns data as pandas.DataFrame object
def get_concentrations_from_file(file_path_concentrations):
    values_concentration = pd.read_excel(file_path_concentrations, sheet_name='Tabelle1')
    return values_concentration

# define model function for linear regression
def linear_model(x, m, c):
    return m * x + c

# define calculation of values with linear regression
def calculation_linear_model(y, m, c):
    return (y-c)/m

# error propagation for linear model
def error_propagation_linear_model(x, m, dm, c, dc, y, dy):
    #Betrag von x multipliziert mit Wurzel aus Summe der relativen Fehler
    return (abs(x) * math.sqrt( (dm / m)**2 + (dc / c)**2 + (dy / y)**2))

# takes data from pandas.DataFrame objects and returns calibration line from first set of data (A1-A12) as a dictionary
def get_calibration(data_object, concentrations_object):
    # get subset with relevant data
    subset = data_object.iloc[2:12, 1:11]
    values_average_deviation = subset.apply(lambda row: (row.mean(), row.std()), axis=1).tolist()
    values_concentration = concentrations_object
    
    x_values = values_concentration.iloc[:9]['konz'].values
    y_values = np.array([y[0] for y in values_average_deviation[:-1]])
    weights = np.array([y[1]**2 for y in values_average_deviation[:-1]])
    value_control = values_average_deviation[-1][0]
    value_control_err = values_average_deviation[-1][1]
    concentration_control = values_concentration.iloc[9:]['konz'].values[0]
    
    # Use curve_fit to perform the weighted linear regression
    popt, pcov = curve_fit(linear_model, x_values, y_values, sigma=weights, absolute_sigma=True)

    # Extract the results: slope, intercept, and their standard errors
    slope, intercept = popt
    slope_err, intercept_err = np.sqrt(np.diag(pcov))

    # Store the results in a dictionary
    calibration_results = {
        'slope': slope,
        'intercept': intercept,
        'slope_err': slope_err,
        'intercept_err': intercept_err
    }

    control_calculated = calculation_linear_model(y = value_control, c = intercept, m = slope)
    control_calculated_err = error_propagation_linear_model(
        x = control_calculated, 
        m = slope, 
        dm = slope_err, 
        c = intercept, 
        dc = intercept_err, 
        y = value_control, 
        dy = value_control_err
    )
    
    calibration_results['control_concentration'] = concentration_control
    calibration_results['control_calculated'] = control_calculated
    calibration_results['control_calculated_err'] = control_calculated_err

    return calibration_results 

# calculate concentrations and return the results as tuple of value and err in a list 
def get_concentration(calibration_object, data_object):

    slope = calibration_object['slope']
    intercept = calibration_object['intercept']
    slope_err = calibration_object['slope_err']
    intercept_err = calibration_object['intercept_err']

    subset = data_object.iloc[14:, 1:11]
    values_average_deviation = subset.apply(lambda row: (row.mean(), row.std()), axis=1).tolist()
    
    #get average and deviation of values in every row in sets of 3
    chunked = [values_average_deviation[i:i+3] for i in range(0, len(values_average_deviation), 3)]
    average_and_errors = []

    for chunk in chunked:
        average = sum(val[0] for val in chunk) / len(chunk)
        sem = math.sqrt(sum(val[1]**2 for val in chunk)) / len(chunk)
        average_and_errors.append((average, sem))

    results_list = []

    for item in average_and_errors:
        new_concentration = calculation_linear_model(y = item[0], c = intercept, m = slope)
        new_concentration_err = error_propagation_linear_model(
            x = new_concentration,
            m = slope,
            dm = slope_err,
            c = intercept,
            dc = intercept_err,
            y = item[0],
            dy = item[1]
        )
        new_tuple = (new_concentration, new_concentration_err)
        results_list.append(new_tuple)

    return results_list

# takes data from calculated concentration and returns amount of substance per area (nmol/cm**2) (a list of list of tuples (average and deviation) )
def get_amounts_per_area(data_list):
    area = 0.8*1.5*2 # cm**2
    volume = 1.1 # ml

    new_list = []
    for item in data_list:
        new_amount = (item[0] * volume) / area * 10**6
        new_amount_err = new_amount * item[1] / item [0]
        new_tuple = (new_amount, new_amount_err)
        new_list.append(new_tuple)
    return new_list

# takes data from calculated concentrations and returns a DataFrame that is structured like a well plate (from a list of list with four tuples) 
def organize_data_well_plate_design(data_list):
    chunked = [data_list[i:i+4] for i in range(0, len(data_list), 4)]
    columns = ['Column 1-3', 'Column 4-6', 'Column 7-9', 'Column 10-12']
    new_data_frame = pd.DataFrame(chunked, columns = columns)

    return new_data_frame

# to make sure the dataframe can be written to an excel-sheet
def remove_tuples_from_df(dataframe):
    #splitting tuples in single columns
    for column in dataframe.columns:
        dataframe[[f'{column}_value', f'{column}_error']] = dataframe[column].apply(pd.Series)
    
    #remove first 4 columns
    new_dataframe = dataframe.iloc[:, 4:].copy()
    return new_dataframe

def make_result_excel_sheet(calibration_data, concentration_data, amount_data, output_path):
    # Create workbook
    workbook = Workbook()
    # Remove the default sheet created by Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    # Write each DataFrame to a specific sheet
    for name, data in [("calibration", calibration_data), ("concentration", concentration_data), ("amount", amount_data)]:
        write_df_to_sheet(workbook, name, data)
    
    # Save workbook
    workbook.save(output_path)

def write_df_to_sheet(workbook, sheet_name, dataframe):
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]
    sheet = workbook.create_sheet(sheet_name)
    for row in dataframe_to_rows(dataframe, index=False, header=True):
        sheet.append(row)

my_data = get_data(file_path_blinds = file_path_blinds, file_path_samples = file_path_samples)
my_concentrations = get_concentrations_from_file(file_path_concentrations = file_path_concentrations)
my_calibration = get_calibration(my_data, my_concentrations)
my_results_concentrations = get_concentration(my_calibration, my_data)
my_results_amounts = get_amounts_per_area(my_results_concentrations)

my_calibration_df = pd.DataFrame(my_calibration, index=[0])
results_as_concentration_df = remove_tuples_from_df(organize_data_well_plate_design(my_results_concentrations))
my_results_amounts_df = remove_tuples_from_df(organize_data_well_plate_design(my_results_amounts))

make_result_excel_sheet(my_calibration_df, results_as_concentration_df, my_results_amounts_df, file_path_output)
print('done')